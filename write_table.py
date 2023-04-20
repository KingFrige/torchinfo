import numpy as np
import pandas as pd
import openpyxl

def write_table(table, filePath, sheetName):
  try:
      # this statement shall raise error for Badzip file
      wb_obj = openpyxl.load_workbook(filename=filePath)
  except:
      # try to create a new file and save at same path
      wb_obj = openpyxl.Workbook()
      wb_obj.save(filePath)
      wb_obj = openpyxl.load_workbook(filename=filePath)
  # 2)创建excel写入对象
  writer = pd.ExcelWriter(filePath, engine="openpyxl", mode='a', if_sheet_exists='replace')
  writer._book = wb_obj
  # 3)写入数据
  data = pd.DataFrame(table)
  data.to_excel(writer, sheet_name=sheetName)
  #writer._save()#不要用这个来保存
  # 4)关闭
  writer.close()
  wb_obj.close()

def main():
  write_table(np.random.randn(10, 5), 'test-param.xlsx', 'summary')

if __name__ == "__main__":
  main()
